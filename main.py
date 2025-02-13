import openpyxl

# Načtení souboru
soubor = './stl.xlsx'  # Zde nahraďte názvem vašeho souboru
workbook = openpyxl.load_workbook(soubor)

# Zadání hodnot uživatelem pro výběr typu průřezu
typ_prurezu = input("Zadejte typ průřezu (např. IPE nebo HEB): ").upper()

# Kontrola, zda list existuje
if typ_prurezu in workbook.sheetnames:
    sheet = workbook[typ_prurezu]  # Vybere správný list podle zadaného typu průřezu

# Zadání velikosti průřezu
velikost_prurezu = input("Zadejte velikost průřezu (např. 80): ")

# Procházení řádků od druhého řádku (první je obvykle hlavička)
for row in range(2, sheet.max_row + 1):
    velikost = sheet.cell(row=row, column=3).value  # Velikost (sloupec C)
    
    # Pokud odpovídá, vypíšeme hodnoty z D a E na stejném řádku
    if str(velikost) == velikost_prurezu:
        hodnota_d = sheet.cell(row=row, column=4).value  # Hodnota ve sloupci D
        hodnota_e = sheet.cell(row=row, column=5).value  # Hodnota ve sloupci E
        print(f"Na řádku {row} nalezeno:")
        print(f"Typ: {typ_prurezu}, Velikost: {velikost}, Hodnota D: {hodnota_d}, Hodnota E: {hodnota_e}")
        break

